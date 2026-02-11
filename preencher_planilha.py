#!/usr/bin/env python3
import argparse
import copy
import re
from io import BytesIO
from pathlib import Path

import pdfplumber
import openpyxl

META_RE = re.compile(r"^META ESPEC[ÍI]FICA\s+(\d+)", re.IGNORECASE)
ITEM_RE = re.compile(
    r"^Item\s*(\d+)\s*(Planejado|Aprovado|Cancelado)?", re.IGNORECASE
)
ACTION_HEADER_KEY = "acao_art"
ACTION_HEADER_NUM_KEY = "acao_art_num"
ACTION_HEADER_PATTERN = re.compile(
    r"^Ação conforme Art\.\s*\d+º\s+da portaria nº 685$",
    re.IGNORECASE,
)
PLAN_SIGNATURE_RE = re.compile(
    r"\b([A-Z]{2})\s*-\s*([A-Z0-9]+)\s*-\s*(20\d{2})\b"
)
ART_PATTERN = re.compile(
    r"^Art\.?\s*(6|7|8)\s*º?\s*(?:\((\d+)\))?\s*:\s*(.*)",
    re.IGNORECASE,
)
ACTION_PATTERN = re.compile(r"^A[cç][aã]o:\s*(.*)", re.IGNORECASE)

CAPTURE_PATTERNS = [
    ("bem", re.compile(r"^(?:Bem|Material)/Servi[cç]o:\s*(.*)", re.IGNORECASE)),
    ("descricao", re.compile(r"^Descri[cç][aã]o:\s*(.*)", re.IGNORECASE)),
    ("destinacao", re.compile(r"^Destina[cç][aã]o:\s*(.*)", re.IGNORECASE)),
    ("unidade", re.compile(r"^Unidade de Medida:\s*(.*)", re.IGNORECASE)),
    ("quantidade", re.compile(r"^Qtd\.?\s*Planejada:\s*(.*)", re.IGNORECASE)),
    ("quantidade", re.compile(r"^Quantidade Planejada:\s*(.*)", re.IGNORECASE)),
    ("natureza", re.compile(r"^Natureza\s*\(ND\):\s*(.*)", re.IGNORECASE)),
    ("instituicao", re.compile(r"^Institui[cç][aã]o:\s*(.*)", re.IGNORECASE)),
    ("valor_total", re.compile(r"^Valor Total:\s*(.*)", re.IGNORECASE)),
]

STOP_PATTERNS = [
    re.compile(r"^C[oó]d\.?\s*Senasp:", re.IGNORECASE),
    re.compile(r"^Valor Origin[aá]rio Planejado:", re.IGNORECASE),
    re.compile(r"^Valor Suplementar Planejado:", re.IGNORECASE),
    re.compile(r"^Valor Rendimento Planejado:", re.IGNORECASE),
]

OUTPUT_HEADERS = [
    "Número da Meta Específica",
    "Número do Item",
    "Ação conforme Art. 7º da portaria nº 685",
    "Material/Serviço",
    "Descrição",
    "Destinação",
    "Instituição",
    "Natureza da Despesa",
    "Quantidade Planejada",
    "Unidade de Medida",
    "Valor Planejado Total",
    "Status do Item",
]

ANALYSIS_TEMPLATE_TITLE = "ANÁLISE DOS ELEMENTOS DO PLANO DE APLICAÇÃO"
ANALYSIS_BLOCK_START_ROW = 14
ANALYSIS_BLOCK_HEIGHT = 11
ANALYSIS_BLOCK_START_COL = 1  # A
ANALYSIS_BLOCK_END_COL = 10  # J


def normalize(text: str) -> str:
    text = re.sub(r"\s+", " ", text or "").strip()
    return text


def blank_if_dash_only(value: str) -> str:
    text = normalize(value)
    if re.fullmatch(r"[-–—]+", text):
        return ""
    return text


def strip_currency(value: str) -> str:
    value = (value or "").replace("R$", "").strip()
    value = value.replace(".", "")
    value = re.sub(r"\s+", "", value)
    return value


def format_currency(value: str) -> str:
    value = strip_currency(value)
    if not value:
        return ""
    if "," in value:
        integer_part, decimal_part = value.split(",", 1)
    else:
        integer_part, decimal_part = value, "00"
    integer_part = re.sub(r"[^0-9]", "", integer_part)
    decimal_part = re.sub(r"[^0-9]", "", decimal_part)[:2].ljust(2, "0")
    integer_part = integer_part.lstrip("0") or "0"
    grouped = ""
    while integer_part:
        grouped = integer_part[-3:] + (f".{grouped}" if grouped else "")
        integer_part = integer_part[:-3]
    return f"R$ {grouped},{decimal_part}"


def parse_int(value: str):
    digits = re.sub(r"[^0-9]", "", value or "")
    return int(digits) if digits else ""


def normalize_pdf_text(text: str) -> str:
    text = text.replace("\x0c", "\n")
    text = re.sub(
        r"(META ESPEC[ÍI]FICA\s+\d+)", r"\n\1\n", text, flags=re.IGNORECASE
    )
    text = re.sub(
        r"(Item\s*\d+\s*(?:Planejado|Aprovado|Cancelado)?)",
        r"\n\1\n",
        text,
        flags=re.IGNORECASE,
    )
    return text


def clean_lines(text: str):
    lines = []
    for raw in text.splitlines():
        line = raw.strip()
        if not line:
            continue
        if re.match(r"^\d{2}/\d{2}/\d{4},", line):
            continue
        if "Planos de Aplicação" in line and re.search(r"\d{2}/\d{2}/\d{4}", line):
            continue
        if line.startswith("https://apps.mj.gov.br/"):
            continue
        lines.append(line)
    return lines


def extract_lines_from_pdf(pdf_path: Path):
    lines = []
    with pdfplumber.open(str(pdf_path)) as pdf:
        for page in pdf.pages:
            text = normalize_pdf_text(page.extract_text() or "")
            lines.extend(clean_lines(text))
    return lines


def extract_lines_from_pdf_file(file_obj):
    file_obj.seek(0)
    lines = []
    with pdfplumber.open(file_obj) as pdf:
        for page in pdf.pages:
            text = normalize_pdf_text(page.extract_text() or "")
            lines.extend(clean_lines(text))
    return lines


def extract_plan_signature(lines):
    max_lines = min(len(lines), 120)
    for idx in range(max_lines):
        line = (lines[idx] or "").strip()
        if not line:
            continue
        match = PLAN_SIGNATURE_RE.search(line.upper())
        if match:
            return {
                "sigla": match.group(2).upper(),
                "ano": int(match.group(3)),
                "raw_line": line,
            }
    return {"sigla": None, "ano": None, "raw_line": None}


def resolve_art_by_plan_rule(sigla, ano):
    if not sigla or not ano:
        return None
    sigla = str(sigla).upper()
    if sigla in {"ECV", "FISPDS", "RMVI"} and 2019 <= ano <= 2025:
        return "6"
    if sigla == "EVM" and 2023 <= ano <= 2025:
        return "7"
    if sigla in {"VPSP", "MQVPSP"} and 2019 <= ano <= 2025:
        return "8"
    return None


def is_analysis_template_sheet(ws) -> bool:
    title = normalize(str(ws["A2"].value or "")).upper()
    return ANALYSIS_TEMPLATE_TITLE in title


def is_analysis_template_file(template_path: Path) -> bool:
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    return is_analysis_template_sheet(ws)


def _header_key(header: str) -> str:
    return ACTION_HEADER_KEY if ACTION_HEADER_PATTERN.match(header) else header


def get_template_header_info_by_row(template_path: Path, header_row: int):
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    return get_header_info_from_ws(ws, header_row)


def get_header_info_from_ws(ws, header_row: int):
    headers = []
    header_map = {}
    for cell in ws[header_row]:
        if cell.value:
            header = str(cell.value).strip()
            headers.append(header)
            key = _header_key(header)
            if key not in header_map:
                header_map[key] = cell.column
    return headers, header_map


def parse_items(lines):
    items = []
    current_meta = None
    current_item = None
    current_status = None
    current_lines = []

    def flush():
        nonlocal current_item, current_lines, current_status
        if current_meta is None or current_item is None:
            return
        items.append({
            "meta": current_meta,
            "item": current_item,
            "status": current_status or "",
            "lines": current_lines[:],
        })
        current_item = None
        current_status = None
        current_lines = []

    for line in lines:
        meta_match = META_RE.match(line)
        if meta_match:
            flush()
            current_meta = int(meta_match.group(1))
            continue
        item_match = ITEM_RE.match(line)
        if item_match:
            flush()
            current_item = int(item_match.group(1))
            current_status = (item_match.group(2) or "").capitalize()
            current_lines = []
            continue
        if current_item is not None:
            current_lines.append(line)

    flush()
    return items


META_GERAL_LINE_RE = re.compile(r"^Meta Geral$", re.IGNORECASE)
INDICADOR_GERAL_LINE_RE = re.compile(r"^Indicador Geral de Resultado$", re.IGNORECASE)
VALOR_REFERENCIA_RE = re.compile(r"valor de refer[eê]ncia\s*:", re.IGNORECASE)
META_ESPECIFICA_LINE_RE = re.compile(r"^META ESPEC[ÍI]FICA\s+(\d+)", re.IGNORECASE)
SECTION_LABEL_PATTERNS = [
    ("descricao_indicador", re.compile(r"^Descri[cç][aã]o do Indicador:\s*(.*)", re.IGNORECASE)),
    ("formula", re.compile(r"^F[oó]rmula:\s*(.*)", re.IGNORECASE)),
    ("carteira_mjsp", re.compile(r"^Carteira de Pol[íi]ticas do MJSP:\s*(.*)", re.IGNORECASE)),
    ("meta_pnsp", re.compile(r"^Meta do PNSP:\s*(.*)", re.IGNORECASE)),
    ("meta_pesp", re.compile(r"^Meta do PESP:\s*(.*)", re.IGNORECASE)),
]
META_PESP_CUTOFF_RE = re.compile(
    r"\b(?:Periodicidade|Fonte(?:/Ano)?|Valor de Refer[eê]ncia(?:/Fonte)?)\s*:",
    re.IGNORECASE,
)


def extract_meta_geral(lines) -> str:
    for idx, line in enumerate(lines):
        if META_GERAL_LINE_RE.match(line):
            collected = []
            for next_line in lines[idx + 1:]:
                if re.match(r"^(Justificativa|Indicador Geral de Resultado|META ESPEC[ÍI]FICA)", next_line, re.IGNORECASE):
                    break
                collected.append(next_line)
            return blank_if_dash_only(" ".join(collected))
    return ""


def extract_indicador_geral_completo(lines) -> str:
    for idx, line in enumerate(lines):
        if not INDICADOR_GERAL_LINE_RE.match(line):
            continue
        collected = []
        for next_line in lines[idx + 1:]:
            if re.match(r"^(Meta Geral|META ESPEC[ÍI]FICA)", next_line, re.IGNORECASE):
                break
            collected.append(next_line)
        return blank_if_dash_only(" ".join(collected))
    return ""


def extract_indicador_geral_valor_referencia(lines) -> str:
    for idx, line in enumerate(lines):
        marker_match = VALOR_REFERENCIA_RE.search(line)
        if not marker_match:
            continue
        collected = [line[marker_match.start():].strip()]
        for next_line in lines[idx + 1:]:
            if re.match(r"^(META ESPEC[ÍI]FICA|Descri[cç][aã]o do Indicador:|Itens da Meta|Status:)", next_line, re.IGNORECASE):
                break
            collected.append(next_line)
        return blank_if_dash_only(" ".join(collected))
    return ""


def extract_analysis_data(lines):
    return {
        "zero_indicador_geral": extract_indicador_geral_completo(lines),
        "one_meta_geral": extract_meta_geral(lines),
        "three_valor_referencia": extract_indicador_geral_valor_referencia(lines),
        "sections": extract_meta_especifica_sections(lines),
    }


def _finalize_meta_section(section):
    result = {"numero_meta": section["numero_meta"]}
    for key in (
        "meta_texto",
        "descricao_indicador",
        "formula",
        "meta_pesp",
        "meta_pnsp",
        "carteira_mjsp",
    ):
        result[key] = blank_if_dash_only(" ".join(section.get(key, [])))
    return result


def _trim_meta_pesp(text: str) -> str:
    text = blank_if_dash_only(text)
    if not text:
        return ""
    cutoff = META_PESP_CUTOFF_RE.search(text)
    if cutoff:
        return blank_if_dash_only(text[:cutoff.start()])
    return blank_if_dash_only(text)


def extract_meta_especifica_sections(lines):
    sections = []
    current = None
    current_field = None

    for line in lines:
        meta_match = META_ESPECIFICA_LINE_RE.match(line)
        if meta_match:
            if current is not None:
                sections.append(_finalize_meta_section(current))
            current = {
                "numero_meta": int(meta_match.group(1)),
                "meta_texto": [],
                "descricao_indicador": [],
                "formula": [],
                "meta_pesp": [],
                "meta_pnsp": [],
                "carteira_mjsp": [],
            }
            current_field = "meta_texto"
            continue

        if current is None:
            continue

        if re.match(r"^Status:", line, re.IGNORECASE):
            current_field = None
            continue
        if re.match(r"^Itens da Meta$", line, re.IGNORECASE):
            current_field = None
            continue
        if ITEM_RE.match(line):
            current_field = None
            continue

        matched_label = False
        for field_key, pattern in SECTION_LABEL_PATTERNS:
            match = pattern.match(line)
            if match:
                current_field = field_key
                content = match.group(1).strip()
                if content:
                    current[field_key].append(content)
                matched_label = True
                break
        if matched_label:
            continue

        if current_field:
            current[current_field].append(line)

    if current is not None:
        sections.append(_finalize_meta_section(current))
    for section in sections:
        section["meta_pesp"] = _trim_meta_pesp(section.get("meta_pesp", ""))
    return sections


def extract_fields(item_lines):
    fields = {key: [] for key, _ in CAPTURE_PATTERNS}
    fields["acao"] = []
    fields["art"] = []
    fields["art_num"] = ""
    current_field = None

    for line in item_lines:
        matched = False
        for stop in STOP_PATTERNS:
            if stop.match(line):
                current_field = None
                matched = True
                break
        if matched:
            continue

        action_match = ACTION_PATTERN.match(line)
        if action_match:
            current_field = "acao"
            action_body = action_match.group(1).strip()
            if action_body:
                fields[current_field].append(action_body)
            continue

        art_match = ART_PATTERN.match(line)
        if art_match:
            current_field = "art"
            art_num = art_match.group(1)
            art_body = art_match.group(3).strip()
            if art_body:
                fields[current_field].append(art_body)
            fields["art_num"] = art_num
            continue

        for field, pattern in CAPTURE_PATTERNS:
            match = pattern.match(line)
            if match:
                current_field = field
                content = match.group(1).strip()
                if content:
                    fields[current_field].append(content)
                matched = True
                break
        if matched:
            continue

        if current_field:
            fields[current_field].append(line)

    for key in fields:
        fields[key] = blank_if_dash_only(" ".join(fields[key]))

    return fields


def _inject_reference_text(base_text: str, reference_text: str) -> str:
    if not reference_text:
        return base_text
    marker = "A referência informada foi:"
    if marker not in base_text:
        return reference_text
    return f"{base_text.split(marker, 1)[0]}{marker}\n\n\n\n{reference_text}"


def _inject_meta_text(base_text: str, marker: str, value: str) -> str:
    if not value:
        return base_text
    if marker not in base_text:
        return value
    before, after = base_text.split(marker, 1)
    suffix_idx = after.find("Existe aderência")
    suffix = f"\n\n\n\n{after[suffix_idx:].strip()}" if suffix_idx != -1 else ""
    return f"{before}{marker}\n\n\n\n{value}{suffix}"


def _inject_descricao_formula(base_text: str, descricao: str, formula: str) -> str:
    if not descricao and not formula:
        return base_text
    marker_desc = "Descrição do Indicador:"
    marker_formula = "Fórmula:"
    if marker_desc not in base_text or marker_formula not in base_text:
        parts = []
        if descricao:
            parts.append(f"Descrição do Indicador: {descricao}")
        if formula:
            parts.append(f"Fórmula: {formula}")
        return "\n\n".join(parts)

    pre = base_text.split(marker_desc, 1)[0]
    after_desc = base_text.split(marker_desc, 1)[1]
    after_formula = after_desc.split(marker_formula, 1)[1] if marker_formula in after_desc else ""
    suffix_idx = after_formula.find("O indicador")
    suffix = f"\n\n{after_formula[suffix_idx:].strip()}" if suffix_idx != -1 else ""
    desc_value = descricao or ""
    formula_value = formula or ""
    return f"{pre}{marker_desc}\n{desc_value}\n\n{marker_formula}\n{formula_value}{suffix}"


def replace_placeholder_segment(base_text: str, token: str, value: str) -> str:
    text = str(base_text or "")
    pattern = re.compile(re.escape(token) + r".*?" + re.escape(token), re.DOTALL)
    if not pattern.search(text):
        return text
    return pattern.sub(value or "", text)


def set_cell_font_black(ws, cell_ref: str):
    cell = ws[cell_ref]
    font = copy.copy(cell.font)
    font.color = "FF000000"
    cell.font = font


def set_row_top_fonts_black(ws, row: int, start_col: int = 1, end_col: int = 10):
    for col in range(start_col, end_col + 1):
        set_cell_font_black(ws, f"{openpyxl.utils.get_column_letter(col)}{row}")


def collect_analysis_missing_cells(analysis_data):
    missing_cells = set()
    if not blank_if_dash_only(analysis_data.get("zero_indicador_geral", "")):
        missing_cells.add("F10")
    if not blank_if_dash_only(analysis_data.get("one_meta_geral", "")):
        missing_cells.add("A8")

    sections = analysis_data.get("sections") or []
    reference = blank_if_dash_only(analysis_data.get("three_valor_referencia", ""))
    for idx, section in enumerate(sections, start=1):
        start_row = ANALYSIS_BLOCK_START_ROW + (idx - 1) * ANALYSIS_BLOCK_HEIGHT
        if not blank_if_dash_only(section.get("meta_texto", "")):
            missing_cells.add(f"A{start_row}")
        if not reference:
            missing_cells.add(f"E{start_row}")
        if not blank_if_dash_only(section.get("descricao_indicador", "")) or not blank_if_dash_only(
            section.get("formula", "")
        ):
            missing_cells.add(f"F{start_row}")
        if not blank_if_dash_only(section.get("meta_pesp", "")):
            missing_cells.add(f"G{start_row}")
        if not blank_if_dash_only(section.get("meta_pnsp", "")):
            missing_cells.add(f"H{start_row}")
        if not blank_if_dash_only(section.get("carteira_mjsp", "")):
            missing_cells.add(f"I{start_row}")
    return sorted(missing_cells)


def build_material(bem, descricao, destinacao):
    parts = []
    if bem:
        parts.append(f"Bem/Serviço: {bem}")
    if descricao:
        parts.append(f"Descrição: {descricao}")
    if destinacao:
        parts.append(f"Destinação: {destinacao}")
    return " | ".join(parts)


def _count_analysis_blocks(ws) -> int:
    count = 0
    for merged in ws.merged_cells.ranges:
        if (
            merged.min_col == ANALYSIS_BLOCK_START_COL
            and merged.max_col == ANALYSIS_BLOCK_START_COL
            and (merged.max_row - merged.min_row + 1) == ANALYSIS_BLOCK_HEIGHT
            and merged.min_row >= ANALYSIS_BLOCK_START_ROW
            and (merged.min_row - ANALYSIS_BLOCK_START_ROW) % ANALYSIS_BLOCK_HEIGHT == 0
        ):
            count += 1
    return max(count, 1)


def _copy_analysis_block(ws, src_start_row: int, dst_start_row: int):
    for row_offset in range(ANALYSIS_BLOCK_HEIGHT):
        src_row = src_start_row + row_offset
        dst_row = dst_start_row + row_offset
        ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
        for col in range(ANALYSIS_BLOCK_START_COL, ANALYSIS_BLOCK_END_COL + 1):
            src_cell = ws.cell(src_row, col)
            dst_cell = ws.cell(dst_row, col)
            dst_cell.value = src_cell.value
            dst_cell.font = copy.copy(src_cell.font)
            dst_cell.fill = copy.copy(src_cell.fill)
            dst_cell.border = copy.copy(src_cell.border)
            dst_cell.alignment = copy.copy(src_cell.alignment)
            dst_cell.number_format = src_cell.number_format
            dst_cell.protection = copy.copy(src_cell.protection)

    shift = dst_start_row - src_start_row
    template_merges = [
        rng
        for rng in list(ws.merged_cells.ranges)
        if (
            rng.min_row >= src_start_row
            and rng.max_row < src_start_row + ANALYSIS_BLOCK_HEIGHT
            and rng.min_col >= ANALYSIS_BLOCK_START_COL
            and rng.max_col <= ANALYSIS_BLOCK_END_COL
        )
    ]
    for rng in template_merges:
        ws.merge_cells(
            start_row=rng.min_row + shift,
            start_column=rng.min_col,
            end_row=rng.max_row + shift,
            end_column=rng.max_col,
        )


def _unmerge_analysis_block_region(ws, block_start_row: int):
    block_end_row = block_start_row + ANALYSIS_BLOCK_HEIGHT - 1
    to_unmerge = [
        str(rng)
        for rng in list(ws.merged_cells.ranges)
        if (
            rng.min_row >= block_start_row
            and rng.max_row <= block_end_row
            and rng.min_col >= ANALYSIS_BLOCK_START_COL
            and rng.max_col <= ANALYSIS_BLOCK_END_COL
        )
    ]
    for rng in to_unmerge:
        ws.unmerge_cells(rng)


def _shift_row_dimensions_on_insert(ws, insert_at: int, amount: int):
    if amount <= 0:
        return
    rows_to_shift = sorted(
        [
            row_idx
            for row_idx in ws.row_dimensions
            if isinstance(row_idx, int) and row_idx >= insert_at
        ],
        reverse=True,
    )
    for row_idx in rows_to_shift:
        dim_copy = copy.copy(ws.row_dimensions[row_idx])
        dim_copy.index = row_idx + amount
        ws.row_dimensions[row_idx + amount] = dim_copy
        del ws.row_dimensions[row_idx]


def _ranges_overlap(a, b) -> bool:
    return not (
        a[2] < b[0]
        or b[2] < a[0]
        or a[3] < b[1]
        or b[3] < a[1]
    )


def _insert_rows_preserving_merges(ws, insert_at: int, amount: int):
    if amount <= 0:
        return
    original_ranges = [
        (rng.min_row, rng.min_col, rng.max_row, rng.max_col)
        for rng in list(ws.merged_cells.ranges)
    ]
    for rng in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(rng))
    ws.insert_rows(insert_at, amount)
    _shift_row_dimensions_on_insert(ws, insert_at, amount)
    rebuilt = []

    def add_range(min_row, min_col, max_row, max_col):
        if min_row > max_row or min_col > max_col:
            return
        if min_row == max_row and min_col == max_col:
            return
        candidate = (min_row, min_col, max_row, max_col)
        for existing in rebuilt:
            if _ranges_overlap(candidate, existing):
                return
        rebuilt.append(candidate)

    for min_row, min_col, max_row, max_col in original_ranges:
        if max_row < insert_at:
            add_range(min_row, min_col, max_row, max_col)
            continue
        if min_row >= insert_at:
            add_range(
                min_row + amount,
                min_col,
                max_row + amount,
                max_col,
            )
            continue
        # Split merges that cross the insertion point to avoid invalid overlaps.
        add_range(min_row, min_col, insert_at - 1, max_col)
        add_range(insert_at + amount, min_col, max_row + amount, max_col)

    for min_row, min_col, max_row, max_col in rebuilt:
        ws.merge_cells(
            start_row=min_row,
            start_column=min_col,
            end_row=max_row,
            end_column=max_col,
        )


def _ensure_analysis_blocks(ws, required_blocks: int):
    existing_blocks = _count_analysis_blocks(ws)
    if required_blocks <= existing_blocks:
        return

    items_title_row = None
    for row in range(1, ws.max_row + 1):
        value = normalize(str(ws.cell(row=row, column=1).value or "")).upper()
        if value == "ITENS DE CONTRATAÇÃO":
            items_title_row = row
            break

    extra_blocks = required_blocks - existing_blocks
    additional_rows_needed = extra_blocks * ANALYSIS_BLOCK_HEIGHT
    insert_at = ANALYSIS_BLOCK_START_ROW + existing_blocks * ANALYSIS_BLOCK_HEIGHT
    minimum_gap_rows = 1
    reusable_gap_rows = 0
    if items_title_row and items_title_row > insert_at:
        reusable_gap_rows = max(0, (items_title_row - insert_at) - minimum_gap_rows)
    rows_to_insert = max(0, additional_rows_needed - reusable_gap_rows)
    _insert_rows_preserving_merges(ws, insert_at, rows_to_insert)
    for block_idx in range(existing_blocks + 1, required_blocks + 1):
        dst_start_row = ANALYSIS_BLOCK_START_ROW + (block_idx - 1) * ANALYSIS_BLOCK_HEIGHT
        _copy_analysis_block(ws, ANALYSIS_BLOCK_START_ROW, dst_start_row)


def fill_analysis_template(ws, lines):
    analysis_data = extract_analysis_data(lines)
    indicador_geral = analysis_data["zero_indicador_geral"]
    meta_geral = analysis_data["one_meta_geral"]
    valor_referencia = analysis_data["three_valor_referencia"]
    sections = analysis_data["sections"]

    base_a8 = str(ws["A8"].value or "")
    a8_replaced = replace_placeholder_segment(base_a8, "1*", meta_geral)
    if a8_replaced != base_a8:
        ws["A8"] = a8_replaced
    elif meta_geral:
        ws["A8"] = meta_geral
    set_cell_font_black(ws, "A8")

    base_f10 = str(ws["F10"].value or "")
    f10_replaced = replace_placeholder_segment(base_f10, "0*", indicador_geral)
    if f10_replaced != base_f10:
        ws["F10"] = f10_replaced
    elif indicador_geral:
        ws["F10"] = indicador_geral
    set_cell_font_black(ws, "F10")

    if not sections:
        return

    _ensure_analysis_blocks(ws, len(sections))

    for idx in range(2, len(sections) + 1):
        start_row = ANALYSIS_BLOCK_START_ROW + (idx - 1) * ANALYSIS_BLOCK_HEIGHT
        # Keep static template text/format across all blocks (B..J included).
        _unmerge_analysis_block_region(ws, start_row)
        _copy_analysis_block(ws, ANALYSIS_BLOCK_START_ROW, start_row)

    for idx, section in enumerate(sections, start=1):
        start_row = ANALYSIS_BLOCK_START_ROW + (idx - 1) * ANALYSIS_BLOCK_HEIGHT
        meta_text_raw = section.get("meta_texto", "")
        meta_text = re.sub(r"^\d+\s*-\s*", "", meta_text_raw).strip()
        two_meta_texto = f"{idx} - {meta_text}" if meta_text else ""

        cell_a = f"A{start_row}"
        base_a = str(ws[cell_a].value or "")
        a_replaced = replace_placeholder_segment(base_a, "2*", two_meta_texto)
        ws[cell_a] = a_replaced if a_replaced != base_a else two_meta_texto

        cell_e = f"E{start_row}"
        base_e = str(ws[cell_e].value or "")
        e_replaced = replace_placeholder_segment(base_e, "3*", valor_referencia)
        if e_replaced == base_e:
            e_replaced = _inject_reference_text(base_e, valor_referencia)
        ws[cell_e] = e_replaced

        cell_f = f"F{start_row}"
        base_f = str(ws[cell_f].value or "")
        f_replaced = replace_placeholder_segment(base_f, "4*", section.get("descricao_indicador", ""))
        f_replaced = replace_placeholder_segment(f_replaced, "5*", section.get("formula", ""))
        if f_replaced == base_f:
            f_replaced = _inject_descricao_formula(
                base_f,
                section.get("descricao_indicador", ""),
                section.get("formula", ""),
            )
        ws[cell_f] = f_replaced

        cell_g = f"G{start_row}"
        base_g = str(ws[cell_g].value or "")
        g_replaced = replace_placeholder_segment(base_g, "6*", section.get("meta_pesp", ""))
        if g_replaced == base_g:
            g_replaced = _inject_meta_text(
                base_g,
                "A Meta informada foi:",
                section.get("meta_pesp", ""),
            )
        ws[cell_g] = g_replaced

        cell_h = f"H{start_row}"
        base_h = str(ws[cell_h].value or "")
        h_replaced = replace_placeholder_segment(base_h, "7*", section.get("meta_pnsp", ""))
        if h_replaced == base_h:
            h_replaced = _inject_meta_text(
                base_h,
                "A Meta informada foi:",
                section.get("meta_pnsp", ""),
            )
        ws[cell_h] = h_replaced

        cell_i = f"I{start_row}"
        base_i = str(ws[cell_i].value or "")
        i_replaced = replace_placeholder_segment(
            base_i, "8*", section.get("carteira_mjsp", "")
        )
        if i_replaced == base_i:
            i_replaced = _inject_meta_text(
                base_i,
                "A política informada foi:",
                section.get("carteira_mjsp", ""),
            )
        ws[cell_i] = i_replaced
        set_row_top_fonts_black(ws, start_row, 1, 10)

def fill_worksheet(ws, rows, header_map, start_row=3):
    # Clear previous data (keep headers)
    max_col = max(header_map.values()) if header_map else ws.max_column
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, max_col=max_col):
        for cell in row:
            cell.value = None

    style_template_row = start_row if start_row <= ws.max_row else None
    style_template_has_custom_style = False
    if style_template_row is not None and header_map:
        style_template_has_custom_style = any(
            ws.cell(style_template_row, col_idx).style_id != 0
            for col_idx in header_map.values()
        )

    for idx, row_data in enumerate(rows, start=start_row):
        if (
            style_template_row is not None
            and style_template_has_custom_style
            and header_map
            and idx != style_template_row
            and all(ws.cell(idx, col_idx).style_id == 0 for col_idx in header_map.values())
        ):
            for col_idx in header_map.values():
                ws.cell(idx, col_idx)._style = copy.copy(
                    ws.cell(style_template_row, col_idx)._style
                )
            template_height = ws.row_dimensions[style_template_row].height
            if template_height is not None:
                ws.row_dimensions[idx].height = template_height
        for header, col_idx in header_map.items():
            ws.cell(row=idx, column=col_idx, value=row_data.get(header, ""))


def get_template_header_info(template_path: Path):
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    headers, header_map = get_header_info_from_ws(ws, 2)
    if not header_map:
        headers = OUTPUT_HEADERS[:]
        header_map = {}
        for idx, header in enumerate(headers):
            key = _header_key(header)
            if key not in header_map:
                header_map[key] = idx + 1
    return headers, header_map


def find_items_table_header_row(ws):
    for row in range(1, ws.max_row + 1):
        value = normalize(str(ws.cell(row=row, column=1).value or "")).upper()
        if value == "ITENS DE CONTRATAÇÃO":
            return row + 1
    return None


def get_analysis_items_header_info(template_path: Path):
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    header_row = find_items_table_header_row(ws)
    if not header_row:
        return None, [], {}
    headers, header_map = get_header_info_from_ws(ws, header_row)
    return header_row, headers, header_map


def update_action_header(ws, rows, header_map, art_num_preferred=None, header_row=2):
    col_idx = header_map.get(ACTION_HEADER_KEY)
    if not col_idx or not rows:
        return
    art_num = art_num_preferred
    if not art_num:
        art_num = rows[0].get(ACTION_HEADER_NUM_KEY)
    if not art_num:
        return
    if str(art_num) not in {"6", "7", "8"}:
        return
    ws.cell(
        row=header_row,
        column=col_idx,
        value=f"Ação conforme Art. {art_num}º da portaria nº 685",
    )


def write_excel(
    template_path: Path,
    output_path: Path,
    rows,
    header_map,
    art_num_preferred=None,
    source_lines=None,
):
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    if is_analysis_template_sheet(ws):
        fill_analysis_template(ws, source_lines or [])
        header_row = find_items_table_header_row(ws)
        if header_row and rows:
            _, items_header_map = get_header_info_from_ws(ws, header_row)
            update_action_header(
                ws,
                rows,
                items_header_map,
                art_num_preferred=art_num_preferred,
                header_row=header_row,
            )
            fill_worksheet(ws, rows, items_header_map, start_row=header_row + 1)
    else:
        update_action_header(ws, rows, header_map, art_num_preferred=art_num_preferred)
        fill_worksheet(ws, rows, header_map)
    ws.sheet_view.topLeftCell = "A1"
    ws.sheet_view.selection[0].activeCell = "A1"
    ws.sheet_view.selection[0].sqref = "A1"
    ws.sheet_view.zoomScale = 100
    wb.save(output_path)


def generate_excel_bytes(
    template_path: Path,
    rows,
    header_map,
    art_num_preferred=None,
    source_lines=None,
) -> bytes:
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    if is_analysis_template_sheet(ws):
        fill_analysis_template(ws, source_lines or [])
        header_row = find_items_table_header_row(ws)
        if header_row and rows:
            _, items_header_map = get_header_info_from_ws(ws, header_row)
            update_action_header(
                ws,
                rows,
                items_header_map,
                art_num_preferred=art_num_preferred,
                header_row=header_row,
            )
            fill_worksheet(ws, rows, items_header_map, start_row=header_row + 1)
    else:
        update_action_header(ws, rows, header_map, art_num_preferred=art_num_preferred)
        fill_worksheet(ws, rows, header_map)
    ws.sheet_view.topLeftCell = "A1"
    ws.sheet_view.selection[0].activeCell = "A1"
    ws.sheet_view.selection[0].sqref = "A1"
    ws.sheet_view.zoomScale = 100
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_rows(parsed_items, header_map):
    has_descricao = "Descrição" in header_map
    has_destinacao = "Destinação" in header_map
    has_quantidade_unidade = "Quantidade/Unidade" in header_map
    has_valor_status = "Valor/Status" in header_map
    has_unidade_col = "Unidade de Medida" in header_map
    has_status_col = "Status do Item" in header_map
    rows = []
    for item in parsed_items:
        fields = extract_fields(item["lines"])
        if has_descricao or has_destinacao:
            material = fields["bem"]
        else:
            material = build_material(fields["bem"], fields["descricao"], fields["destinacao"])
        valor_total = format_currency(fields["valor_total"])
        quantidade = parse_int(fields["quantidade"])
        unidade = fields["unidade"]
        status_item = item.get("status") or "Planejado"
        quantidade_unidade = ""
        if has_quantidade_unidade and not has_unidade_col:
            if quantidade != "" and unidade:
                quantidade_unidade = f"{quantidade} {unidade}"
            elif quantidade != "":
                quantidade_unidade = str(quantidade)
            elif unidade:
                quantidade_unidade = unidade
        valor_status = ""
        if has_valor_status and not has_status_col:
            if valor_total and status_item:
                valor_status = f"{valor_total} | {status_item}"
            elif valor_total:
                valor_status = valor_total
            elif status_item:
                valor_status = status_item
        row = {
            "Número da Meta Específica": item["meta"],
            "Número do Item": item["item"],
            ACTION_HEADER_KEY: fields["acao"] or fields["art"],
            ACTION_HEADER_NUM_KEY: fields["art_num"],
            "Material/Serviço": material,
            "Descrição": fields["descricao"] if has_descricao else "",
            "Destinação": fields["destinacao"] if has_destinacao else "",
            "Instituição": fields["instituicao"],
            "Natureza da Despesa": fields["natureza"],
            "Quantidade Planejada": quantidade,
            "Unidade de Medida": fields["unidade"],
            "Quantidade/Unidade": quantidade_unidade,
            "Valor Planejado Total": valor_total,
            "Status do Item": status_item,
            "Valor/Status": valor_status,
        }
        rows.append(row)
    return rows


def main():
    parser = argparse.ArgumentParser(description="Preenche planilha a partir do PDF.")
    parser.add_argument("--pdf", default="Planos de Aplicação.pdf", help="PDF de entrada")
    parser.add_argument("--xlsx", default="Itens NT.xlsx", help="Planilha modelo")
    parser.add_argument("--output", default="Itens NT - preenchido.xlsx", help="Planilha de saída")
    args = parser.parse_args()

    pdf_path = Path(args.pdf)
    xlsx_path = Path(args.xlsx)
    output_path = Path(args.output)

    if not pdf_path.exists():
        raise SystemExit(f"PDF não encontrado: {pdf_path}")
    if not xlsx_path.exists():
        raise SystemExit(f"Planilha não encontrada: {xlsx_path}")

    lines = extract_lines_from_pdf(pdf_path)
    parsed_items = parse_items(lines)
    if not parsed_items:
        raise SystemExit("Nenhum item encontrado no PDF.")

    signature = extract_plan_signature(lines)
    art_num_preferred = resolve_art_by_plan_rule(signature["sigla"], signature["ano"])
    _, header_map = get_template_header_info(xlsx_path)
    rows = build_rows(parsed_items, header_map)
    write_excel(
        xlsx_path,
        output_path,
        rows,
        header_map,
        art_num_preferred=art_num_preferred,
        source_lines=lines,
    )

    print(f"Itens extraídos: {len(rows)}")
    print(f"Arquivo gerado: {output_path}")


if __name__ == "__main__":
    main()
