import base64
import hashlib
import json
from io import BytesIO
from pathlib import Path
import urllib.parse
import urllib.request

import streamlit as st
from openpyxl.utils import get_column_letter
import openpyxl

from planilha_engine import (
    extract_lines_from_pdf_file,
    extract_plan_signature,
    resolve_art_by_plan_rule,
    extract_analysis_data,
    collect_analysis_missing_cells,
    is_analysis_template_file,
    get_analysis_items_header_info,
    find_items_table_header_row,
    parse_items,
    build_rows,
    generate_excel_bytes,
    get_template_header_info,
)

BASE_DIR = Path(__file__).resolve().parent
LOCAL_TEMPLATE_PATH = BASE_DIR / "Planilha Base.xlsx"
LOGO_PATH = BASE_DIR / "Logo.png"
TEMPLATE_CACHE_DIR = Path("/tmp/preenche_planilhas")


def _get_secret(name: str, default=None):
    try:
        if name in st.secrets:
            return st.secrets[name]
    except Exception:
        return default
    return default


@st.cache_data(show_spinner=False, ttl=3600)
def _download_template_bytes_from_github(
    repo: str, template_path: str, ref: str, token: str
) -> bytes:
    encoded_path = urllib.parse.quote(template_path, safe="/")
    encoded_ref = urllib.parse.quote(ref, safe="")
    url = f"https://api.github.com/repos/{repo}/contents/{encoded_path}?ref={encoded_ref}"
    request = urllib.request.Request(url)
    request.add_header("Accept", "application/vnd.github+json")
    request.add_header("Authorization", f"Bearer {token}")
    request.add_header("X-GitHub-Api-Version", "2022-11-28")

    with urllib.request.urlopen(request, timeout=20) as response:
        payload = json.loads(response.read().decode("utf-8"))

    content = payload.get("content")
    if not content:
        raise RuntimeError("Resposta do GitHub sem conteúdo da planilha.")
    return base64.b64decode(content)


def resolve_template_path():
    repo = _get_secret("TEMPLATE_GITHUB_REPO") or _get_secret("TEMPLATE_REPO")
    template_path = _get_secret("TEMPLATE_GITHUB_PATH") or _get_secret("TEMPLATE_PATH")
    ref = _get_secret("TEMPLATE_GITHUB_REF") or _get_secret("TEMPLATE_REF") or "main"
    token = _get_secret("GITHUB_TOKEN") or _get_secret("TEMPLATE_GITHUB_TOKEN")

    if repo or template_path:
        if not repo:
            return None, "Defina TEMPLATE_GITHUB_REPO no st.secrets."
        if not token:
            return None, "Defina GITHUB_TOKEN (ou TEMPLATE_GITHUB_TOKEN) no st.secrets."
        resolved_template_path = template_path or "Planilha Base.xlsx"
        try:
            template_bytes = _download_template_bytes_from_github(
                repo=repo,
                template_path=resolved_template_path,
                ref=ref,
                token=token,
            )
            TEMPLATE_CACHE_DIR.mkdir(parents=True, exist_ok=True)
            cache_key = hashlib.sha256(
                f"{repo}|{resolved_template_path}|{ref}".encode("utf-8")
            ).hexdigest()[:12]
            cached_template = TEMPLATE_CACHE_DIR / f"template_{cache_key}.xlsx"
            cached_template.write_bytes(template_bytes)
            return cached_template, None
        except Exception as exc:
            return None, f"Falha ao baixar Planilha Base do GitHub privado: {exc}"

    if LOCAL_TEMPLATE_PATH.exists():
        return LOCAL_TEMPLATE_PATH, None
    return None, "Planilha modelo não encontrada no servidor."

st.set_page_config(page_title="Preenche Planilhas", page_icon="📄", layout="centered")

st.markdown(
    """
    <style>
    .header {
      display: flex;
      align-items: center;
      gap: 16px;
    }
    .header-title {
      font-size: 1.6rem !important;
      font-weight: 600;
      margin: 0;
      line-height: 1.2;
    }
    .logo-wrap {
      width: 64px;
      height: 64px;
      border-radius: 16px;
      overflow: hidden;
      border: 1px solid #e6e6e6;
      flex: 0 0 auto;
    }
    .logo-wrap img {
      width: 64px;
      height: 64px;
      object-fit: cover;
      display: block;
    }
    .brand-bar {
      display: grid;
      grid-template-columns: repeat(5, 1fr);
      height: 6px;
      border-radius: 999px;
      overflow: hidden;
      border: 1px solid #d8dbe0;
      margin-top: 0.6rem;
      margin-bottom: 14px;
      width: 699px;
    }
    .brand-bar span:nth-child(1) { background: #00b140; }
    .brand-bar span:nth-child(2) { background: #ff1b14; }
    .brand-bar span:nth-child(3) { background: #ffd200; }
    .brand-bar span:nth-child(4) { background: #1f4bff; }
    .brand-bar span:nth-child(5) { background: #ff1b14; }
    .app-subtitle { 
      margin: 0 0 16px 0;
    }
    div[data-testid="stDownloadButton"] button {
      background: #217346;
      border: 1px solid #1e6a40;
      color: #ffffff;
    }
    div[data-testid="stDownloadButton"] button:hover {
      background: #1b5e38;
      border-color: #1b5e38;
      color: #ffffff;
    }
    div[data-testid="stDownloadButton"] button:active,
    div[data-testid="stDownloadButton"] button:focus,
    div[data-testid="stDownloadButton"] button:focus-visible {
      background: #1b5e38;
      border-color: #1b5e38;
      color: #ffffff;
      box-shadow: none;
      outline: none;
    }
    .blank-cells {
      border-collapse: collapse;
      width: auto;
    }
    .blank-cells th,
    .blank-cells td {
      border: 1px solid #e6e6e6;
      padding: 8px 12px;
      text-align: left;
    }
    .blank-cells th {
      background: #fafafa;
      font-weight: 600;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

logo_b64 = ""
if LOGO_PATH.exists():
    logo_bytes = LOGO_PATH.read_bytes()
    logo_b64 = base64.b64encode(logo_bytes).decode("ascii")

logo_html = ""
if logo_b64:
    logo_html = f"""
      <div class="logo-wrap">
        <img src="data:image/png;base64,{logo_b64}" alt="Logo" />
      </div>
    """

st.markdown(
    f"""
    <div class="header">
      {logo_html}
      <h1 class="header-title">Gerador de Planilha de Itens - MJSP</h1>
    </div>
    """,
    unsafe_allow_html=True,
)

st.markdown(
    """
    <div class="brand-bar">
      <span></span><span></span><span></span><span></span><span></span>
    </div>
    """,
    unsafe_allow_html=True,
)

st.markdown(
    '<p class="app-subtitle">Faça upload do PDF do Plano de Aplicação e gere a planilha preenchida automaticamente.</p>',
    unsafe_allow_html=True,
)

uploaded_file = st.file_uploader("PDF do Plano", type=["pdf"])

if "result" not in st.session_state:
    st.session_state.result = None

if st.button("Processar", type="primary", disabled=uploaded_file is None):
    template_source, template_error = resolve_template_path()
    if template_error:
        st.error(template_error)
    elif not template_source or not template_source.exists():
        st.error("Planilha modelo não encontrada no servidor.")
    else:
        try:
            with st.status("Processando PDF...", expanded=True) as status:
                status.write("Lendo PDF")
                lines = extract_lines_from_pdf_file(uploaded_file)
                analysis_mode = is_analysis_template_file(template_source)

                status.write("Extraindo itens")
                parsed_items = parse_items(lines)
                if not analysis_mode and not parsed_items:
                    status.update(label="Nenhum item encontrado.", state="error")
                    st.error("Nenhum item encontrado no PDF.")
                    st.session_state.result = None
                else:
                    status.write("Montando planilha")
                    signature = extract_plan_signature(lines)
                    art_num_preferred = resolve_art_by_plan_rule(
                        signature["sigla"], signature["ano"]
                    )
                    if analysis_mode:
                        analysis_data = extract_analysis_data(lines)
                        sections = analysis_data.get("sections", [])
                        header_row, _, items_header_map = get_analysis_items_header_info(
                            template_source
                        )
                        rows = build_rows(parsed_items, items_header_map)
                        excel_bytes = generate_excel_bytes(
                            template_source,
                            rows=rows,
                            header_map={},
                            art_num_preferred=art_num_preferred,
                            source_lines=lines,
                        )
                        missing_cells = set(collect_analysis_missing_cells(analysis_data))
                        missing_rows = set()
                        generated_wb = openpyxl.load_workbook(BytesIO(excel_bytes))
                        generated_ws = generated_wb.active
                        generated_header_row = find_items_table_header_row(generated_ws)
                        start_row = (
                            generated_header_row + 1
                            if generated_header_row
                            else (header_row + 1 if header_row else 3)
                        )
                        for index, row_data in enumerate(rows):
                            excel_row = start_row + index
                            for header, col_index in items_header_map.items():
                                value = row_data.get(header)
                                if value is None or value == "":
                                    cell = f"{get_column_letter(col_index)}{excel_row}"
                                    missing_cells.add(cell)
                                    missing_rows.add(excel_row)
                        st.session_state.result = {
                            "mode": "analysis",
                            "rows": rows,
                            "excel_bytes": excel_bytes,
                            "meta_counts": {s["numero_meta"]: 1 for s in sections},
                            "missing_cells": sorted(missing_cells),
                            "missing_items_count": len(missing_cells),
                            "sections_count": len(sections),
                            "items_count": len(parsed_items),
                        }
                    else:
                        _, header_map = get_template_header_info(template_source)
                        rows = build_rows(parsed_items, header_map)
                        excel_bytes = generate_excel_bytes(
                            template_source,
                            rows,
                            header_map,
                            art_num_preferred=art_num_preferred,
                            source_lines=lines,
                        )

                        meta_counts = {}
                        missing_cells = set()
                        missing_rows = set()
                        start_row = 3
                        for index, row_data in enumerate(rows):
                            meta = row_data.get("Número da Meta Específica")
                            item_num = row_data.get("Número do Item")
                            meta_counts[meta] = meta_counts.get(meta, 0) + 1
                            excel_row = start_row + index
                            for header, col_index in header_map.items():
                                value = row_data.get(header)
                                if value is None or value == "":
                                    cell = f"{get_column_letter(col_index)}{excel_row}"
                                    missing_cells.add(cell)
                                    missing_rows.add(excel_row)

                        st.session_state.result = {
                            "mode": "items",
                            "rows": rows,
                            "excel_bytes": excel_bytes,
                            "meta_counts": meta_counts,
                            "missing_cells": sorted(missing_cells),
                            "missing_items_count": len(missing_rows),
                        }
                    status.update(label="Processamento concluído.", state="complete")
        except Exception as exc:
            st.exception(exc)

result = st.session_state.result
if result:
    mode = result.get("mode", "items")
    total_items = len(result["rows"])
    total_metas = len(result["meta_counts"])
    missing_count = result["missing_items_count"]
    missing_cells = result["missing_cells"]

    st.subheader("Resumo")
    summary_cols = st.columns(3)
    if mode == "analysis":
        summary_cols[0].metric("Metas encontradas", total_metas)
        summary_cols[1].metric("Itens extraídos", result.get("items_count", 0))
        summary_cols[2].metric("Células em branco", len(missing_cells))
    else:
        summary_cols[0].metric("Itens extraídos", total_items)
        summary_cols[1].metric("Metas encontradas", total_metas)
        summary_cols[2].metric("Itens com campos faltantes", missing_count)

    if missing_count:
        st.warning("Alguns itens possuem campos em branco. Veja os detalhes abaixo.")

    st.download_button(
        "Baixar Planilha",
        data=result["excel_bytes"],
        file_name="Planilha de Itens.xlsx",
        mime=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
    )

    if missing_cells:
        st.subheader("Células em branco")
        rows_html = "".join(f"<tr><td>{cell}</td></tr>" for cell in missing_cells)
        st.markdown(
            f"""
            <table class="blank-cells">
              <thead><tr><th>Célula</th></tr></thead>
              <tbody>{rows_html}</tbody>
            </table>
            """,
            unsafe_allow_html=True,
        )

    # Preview e detalhes removidos conforme solicitado.
