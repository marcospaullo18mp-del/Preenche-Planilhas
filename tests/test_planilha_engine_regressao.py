import unittest
from pathlib import Path

import openpyxl

from planilha_engine import (
    ACTION_HEADER_KEY,
    ACTION_HEADER_NUM_KEY,
    extract_fields,
    extract_indicador_geral_completo,
    extract_meta_especifica_sections,
    fill_analysis_template,
    resolve_action_header_title_by_plan,
    resolve_art_by_plan_rule,
    update_action_header,
)


class TestRegressaoVariacoesPDF(unittest.TestCase):
    def test_resolve_action_header_title_by_plan_para_2023_2024(self):
        self.assertEqual(
            resolve_action_header_title_by_plan("RMV", 2023),
            "Ação conforme Art. 5º da portaria nº 439 ou Art. 6º da portaria nº 685",
        )
        self.assertEqual(
            resolve_action_header_title_by_plan("RMVI", 2023),
            "Ação conforme Art. 5º da portaria nº 439 ou Art. 6º da portaria nº 685",
        )
        self.assertEqual(
            resolve_action_header_title_by_plan("EVM", 2024),
            "Ação conforme Art. 6º da portaria nº 439 ou Art. 7º da portaria nº 685",
        )
        self.assertEqual(
            resolve_action_header_title_by_plan("MQVPSP", 2024),
            "Ação conforme Art. 7º da portaria nº 439 ou Art. 8º da portaria nº 685",
        )
        self.assertIsNone(resolve_action_header_title_by_plan("RMV", 2025))

    def test_update_action_header_deve_priorizar_titulo_preferido(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["C2"] = "Ação conforme Art. 7º da portaria nº 685"
        rows = [{ACTION_HEADER_NUM_KEY: "6"}]
        update_action_header(
            ws,
            rows,
            {ACTION_HEADER_KEY: 3},
            art_num_preferred=None,
            action_header_title_preferred=(
                "Ação conforme Art. 5º da portaria nº 439 "
                "ou Art. 6º da portaria nº 685"
            ),
            header_row=2,
        )
        self.assertEqual(
            ws["C2"].value,
            "Ação conforme Art. 5º da portaria nº 439 ou Art. 6º da portaria nº 685",
        )

    def test_update_action_header_deve_usar_artigo_mais_frequente_nas_linhas(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["C2"] = "Ação conforme Art. 7º da portaria nº 685"
        rows = [
            {ACTION_HEADER_NUM_KEY: "8"},
            {ACTION_HEADER_NUM_KEY: "6"},
            {ACTION_HEADER_NUM_KEY: "6"},
            {ACTION_HEADER_NUM_KEY: "5"},
            {ACTION_HEADER_NUM_KEY: ""},
        ]
        update_action_header(ws, rows, {ACTION_HEADER_KEY: 3}, header_row=2)
        self.assertEqual(
            ws["C2"].value,
            "Ação conforme Art. 6º da portaria nº 685",
        )

    def test_resolve_art_by_plan_rule_deve_nao_fixar_artigo_nos_anos_especificos(self):
        self.assertIsNone(resolve_art_by_plan_rule("MQVPSP", 2019))
        self.assertIsNone(resolve_art_by_plan_rule("MQVPSP", 2022))
        self.assertIsNone(resolve_art_by_plan_rule("MQVPSP", 2025))
        self.assertIsNone(resolve_art_by_plan_rule("MQVPSP", 2026))

    def test_extract_fields_deve_capturar_artigo_fora_de_6_7_8(self):
        lines = [
            "Art. 5º (439): RMV | IV - Fortalecimento da capacidade de investigação",
            "Bem/Serviço: Exemplo",
        ]
        fields = extract_fields(lines)
        self.assertEqual(fields["art_num"], "5")
        self.assertEqual(
            fields["art"],
            "RMV | IV - Fortalecimento da capacidade de investigação",
        )

    def test_indicador_formato_antigo_em_linhas_separadas(self):
        lines = [
            "Alguma linha",
            "Indicador Geral de Resultado",
            "Taxa de redução de crimes por 100 mil.",
            "META ESPECÍFICA 1",
        ]
        self.assertEqual(
            extract_indicador_geral_completo(lines),
            "Taxa de redução de crimes por 100 mil.",
        )

    def test_indicador_inline_apos_titulo(self):
        lines = [
            "Indicador Geral de Resultado: Taxa Y",
            "META ESPECÍFICA 1",
        ]
        self.assertEqual(extract_indicador_geral_completo(lines), "Taxa Y")

    def test_indicador_inline_dentro_de_meta_geral(self):
        lines = [
            "Meta Geral: Reduzir índice X Indicador Geral de Resultado: Taxa Z",
            "META ESPECÍFICA 1",
        ]
        self.assertEqual(extract_indicador_geral_completo(lines), "Taxa Z")

    def test_periodicidade_com_fonte_ano_inline_padrao_antigo(self):
        lines = [
            "META ESPECÍFICA 1",
            "Reduzir a violência",
            "Status: Planejado",
            "Descrição do Indicador: Taxa do EVM",
            "Fórmula: (A-B)/A",
            "Carteira de Políticas do MJSP: Política X",
            "Meta do PNSP: Meta PNSP X",
            "Meta do PESP: Meta PESP X",
            "Periodicidade: Anual | Fonte/Ano: SINESP/SENASP/2025",
        ]
        sections = extract_meta_especifica_sections(lines)
        self.assertEqual(len(sections), 1)
        self.assertEqual(sections[0]["periodicidade"], "Anual")
        self.assertEqual(sections[0]["fonte_ano"], "SINESP/SENASP/2025")

    def test_periodicidade_com_valor_referencia_fonte_inline_variacao_nova(self):
        lines = [
            "META ESPECÍFICA 1",
            "Reduzir a violência",
            "Status: Planejado",
            "Descrição do Indicador: Taxa do EVM",
            "Fórmula: (A-B)/A",
            "Carteira de Políticas do MJSP: Política X",
            "Meta do PNSP: Meta PNSP X",
            "Meta do PESP: Meta PESP X",
            "Periodicidade: Anual | Valor de Referência/Fonte: Estupro: 1.956/SINESP/SENASP/2025",
        ]
        sections = extract_meta_especifica_sections(lines)
        self.assertEqual(len(sections), 1)
        self.assertEqual(sections[0]["periodicidade"], "Anual")
        self.assertEqual(
            sections[0]["fonte_ano"],
            "Estupro: 1.956/SINESP/SENASP/2025",
        )

    def test_bloco_analise_compacto_nao_deve_inserir_linhas_vazias_entre_metas(self):
        wb = openpyxl.load_workbook(
            Path(__file__).resolve().parents[1] / "Planilha Base(atualizada).xlsx"
        )
        ws = wb.active

        lines = [
            "Meta Geral",
            "Texto da meta geral",
            "Indicador Geral de Resultado",
            "Texto do indicador geral",
            "META ESPECÍFICA 1",
            "Meta 1",
            "Status: Planejado",
            "Descrição do Indicador: d1",
            "Fórmula: f1",
            "Carteira de Políticas do MJSP: c1",
            "Meta do PNSP: p1",
            "Meta do PESP: e1",
            "META ESPECÍFICA 2",
            "Meta 2",
            "Status: Planejado",
            "Descrição do Indicador: d2",
            "Fórmula: f2",
            "Carteira de Políticas do MJSP: c2",
            "Meta do PNSP: p2",
            "Meta do PESP: e2",
            "META ESPECÍFICA 3",
            "Meta 3",
            "Status: Planejado",
            "Descrição do Indicador: d3",
            "Fórmula: f3",
            "Carteira de Políticas do MJSP: c3",
            "Meta do PNSP: p3",
            "Meta do PESP: e3",
            "META ESPECÍFICA 4",
            "Meta 4",
            "Status: Planejado",
            "Descrição do Indicador: d4",
            "Fórmula: f4",
            "Carteira de Políticas do MJSP: c4",
            "Meta do PNSP: p4",
            "Meta do PESP: e4",
        ]

        fill_analysis_template(ws, lines)

        self.assertTrue(str(ws["A14"].value).startswith("1 - "))
        self.assertTrue(str(ws["A15"].value).startswith("2 - "))
        self.assertTrue(str(ws["A16"].value).startswith("3 - "))
        self.assertTrue(str(ws["A17"].value).startswith("4 - "))
        self.assertEqual(ws["A18"].value, None)
        self.assertEqual(ws["A19"].value, "ITENS DE CONTRATAÇÃO")


if __name__ == "__main__":
    unittest.main()
